import argparse
import json
import os
import re
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = os.path.join(os.path.expanduser("~"), ".mytime-booker", "timecard_template.xlsx")
DEFAULT_OUTPUT = os.path.join(os.path.expanduser("~"), "Downloads", "timecard_output.xlsx")

TYPE_DEFAULT = "Normal -AT"


def load_events_from_json(json_path):
    with open(json_path) as f:
        return json.load(f)


def build_booking_rows(events):
    """
    Convert pre-mapped calendar events (produced by the agent in Phase 2) into
    timecard rows. Each event is expected to already carry project_id, project_name,
    task_id, and task_name. Events missing these fields are written as UNMAPPED
    (empty strings) so they still appear in the output for manual follow-up.
    """
    rows = []
    for ev in events:
        date_str = ev.get("date", "")
        if isinstance(date_str, str) and "T" in date_str:
            date_str = date_str.split("T")[0]
        start_time = ev.get("start", "09:00")
        end_time = ev.get("end", "10:00")

        try:
            h = (datetime.strptime(end_time, "%H:%M") - datetime.strptime(start_time, "%H:%M")).seconds / 3600
        except Exception:
            h = 1.0

        # Extract project number and strip it from project_name
        # e.g. "295189 - BAW-CC-Cloud-OU216" → project_number="295189", project_name="BAW-CC-Cloud-OU216"
        project_name = ev.get("project_name", "")
        project_number = str(ev.get("project_id", ""))
        pm = re.match(r"^(\d+)\s*-\s*(.+)$", project_name)
        if pm:
            project_number = pm.group(1)
            project_name = pm.group(2)

        # Extract task number and strip it from task_name — always, regardless of task_id field
        # e.g. "61.2 - Cloud support" → task_number="61.2", task_name="Cloud support"
        task_name_full = ev.get("task_name", "")
        task_number = str(ev.get("task_id", ""))
        task_name_out = task_name_full
        m = re.match(r"^(\S+)\s*-\s*(.+)$", task_name_full)
        if m:
            task_number = m.group(1)
            task_name_out = m.group(2)

        rows.append({
            "project_id": project_number,
            "project_name": project_name,
            "task_id": task_number,
            "task_name": task_name_out,
            "type": TYPE_DEFAULT,
            "date": date_str,
            "hours": str(round(h, 2)),
            "comment": ev.get("title", "Meeting"),
            "time_from": start_time,
            "time_to": end_time,
        })
    return rows


def save_timecard(rows, output_path, template_path=None):
    tp = template_path or TEMPLATE_PATH
    if os.path.exists(tp):
        wb = openpyxl.load_workbook(tp)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timecard"
        headers = ["Project number", "Project name", "Task number", "Task name",
                    "Type", "Date", "Hours", "Comment", "Time from", "Time to"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, size=13, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="000000")
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

    ws = wb["Timecard"]
    for row_data in rows:
        ws.append([
            row_data["project_id"],
            row_data["project_name"],
            row_data["task_id"],
            row_data["task_name"],
            row_data["type"],
            row_data["date"],
            row_data["hours"],
            row_data["comment"],
            row_data["time_from"],
            row_data["time_to"],
        ])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"[book-timecard] Saved {len(rows)} row(s) to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Write pre-mapped calendar events into a MyTime timecard xlsx")
    parser.add_argument("--events", required=True, help="Path to JSON file with confirmed, pre-mapped calendar events")
    parser.add_argument("--template", default=TEMPLATE_PATH, help="Path to the blank template xlsx")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output path for the filled xlsx")
    args = parser.parse_args()

    if not os.path.exists(args.events):
        print(f"[book-timecard] ERROR: events file not found: {args.events}")
        sys.exit(1)

    events = load_events_from_json(args.events)
    print(f"[book-timecard] Loaded {len(events)} events")

    rows = build_booking_rows(events)

    if not rows:
        print("[book-timecard] No booking rows generated.")
        sys.exit(1)

    for i, r in enumerate(rows, 1):
        status = "OK" if r["project_id"] else "UNMAPPED"
        print(f"  {i}. [{status}] {r['date']} {r['time_from']}-{r['time_to']} | "
              f"{r['project_id']}/{r['task_id']} | {r['comment'][:50]}")

    save_timecard(rows, args.output, args.template)
    print(f"[book-timecard] Done: {args.output}")


if __name__ == "__main__":
    main()
